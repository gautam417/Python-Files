# lab 2: Analyze and plot scores
# Names: Gautam Mehta, Justin Bloem
import matplotlib.pyplot as plt
import csv
import numpy as np
import openpyxl

def readExcel(fileList) :
    for filename in fileList:
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            listO=[]
            listH=[]
            counter = 0
            index=0 
            excelData= ()
            for counter in range(sheet.max_row):
                rowlist1=[]
                rowlist2=[]
                if (sheet.cell(row = counter+1 ,column = 1).value) == "H":
                    listH.append(rowlist1)
                    for index in range(1, sheet.max_column) :
                        rowlist1.append(sheet.cell(row = counter+1 ,column = index+ 1).value)
                elif (sheet.cell(row = counter+1 ,column = 1).value) == "O":
                    listO.append(rowlist2)
                    for index in range(1, sheet.max_column) :
                        rowlist2.append(sheet.cell(row = counter+1 ,column = index+ 1).value)  
            excelData = (listH,listO)
            return (excelData)
        
def readCSV(fileList) :
    csvH=[]
    csvO=[]
    csvData= ()    
    for filename in fileList:
        if filename.endswith(".csv"):
            with open (filename) as fh:
                reader=csv.reader(fh)
                for row in reader:
                    if (row[0] == 'H'):
                        csvH.append(row[1:])
                    elif (row[0] == 'O'):
                        csvO.append(row[1:])
    csvData = (csvH, csvO)
    return (csvData)

def readFile(*fileList) :
    lists= readCSV(fileList)
    listO = []
    listH= []
    for i in lists[0]:
        listH.append(i)
    for i in lists[1]:
        listO.append(i)
        
    lists= readExcel(fileList)
    for i in lists[0]:
        listH.append(i)
    for i in lists[1]:
        listO.append(i)
    arrH=np.array(listH).astype(np.float)
    arrO=np.array(listO).astype(np.float)
    return arrO,arrH
    
def analyze(arrH, arrO):
    percentGradeH = list((arrH.sum(1)/519)*100)
    percentGradeO = list((arrO.sum(1)/519)*100)

    numHyStudentsPassing = 0
    for i in range(len(percentGradeH)):
        float(percentGradeH[i])
        if percentGradeH[i] >= 70:
            numHyStudentsPassing+=1
    successRateHy= numHyStudentsPassing/len(arrH)
    
    numOnStudentsPassing = 0
    for i in range(len(percentGradeO)):
        float(percentGradeO[i])
        if percentGradeO[i] >= 70:
            numOnStudentsPassing+=1      
    successRateOn= numOnStudentsPassing/len(arrO)
    print ("The success rate of students taking Hybrid classes is: ", successRateHy)
    print ("The success rate of students taking Online classes is: ", successRateOn) 
    
    #Histogram
    plt.title("Percentage Grades")
    plt.xlabel("Bins")
    plt.ylabel("Frequency")
    plt.hist((percentGradeH, percentGradeO), color=("blue", "red"), label=("Hybrid","Online"), alpha=0.5, bins=10)
    plt.legend(loc="best")
    plt.show()
    
    #Line plot
    #subsetArrH= np.array(arrH[0:,0:9])
    #subsetArrO= np.array(arrO[0:,0:9])
    #HybridMean= subsetArrH.mean(0)
    #OnlineMean= subsetArrO.mean(0)
    #days = np.arange(1,31)
    #plt.title("Mean Score Data")
    #plt.plot(days, HybridMean)
    #plt.plot(days, OnlineMean)
    #plt.xlabel("Days (t)")
    #plt.ylabel("Mean") 
    #plt.legend(loc="best") 
    #plt.show()
    ##Wasnt able to get the right dimensions for subsetArr
    
def main() :
    (arrH, arrO)= readFile('data1.xlsx','data2.csv','data3.csv')
    analyze(arrH, arrO)

main()